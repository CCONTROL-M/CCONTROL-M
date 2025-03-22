"""
Sistema de exportação de relatórios para CCONTROL-M.

Este módulo implementa funcionalidades para geração e exportação de relatórios
em diferentes formatos (PDF e Excel) com suporte a templates e formatação.
"""
import os
import io
import json
import datetime
from typing import List, Dict, Any, Optional, Union
from pathlib import Path

import pandas as pd
import pdfkit
import jinja2
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.core.config import settings
from app.core.pdf_config import get_pdf_options, configure_pdfkit
from app.utils.logging_config import get_logger

# Configurar logger
logger = get_logger(__name__)

class ReportExporter:
    """Classe para exportação de relatórios em diferentes formatos."""
    
    def __init__(
        self,
        template_dir: str = "templates/reports",
        output_dir: str = "reports",
        company_logo: Optional[str] = None,
        default_font: str = "Arial",
    ):
        """
        Inicializa o exportador de relatórios.
        
        Args:
            template_dir: Diretório com templates de relatório
            output_dir: Diretório para salvar relatórios gerados
            company_logo: Caminho para o logo da empresa
            default_font: Fonte padrão para relatórios
        """
        self.template_dir = os.path.join(settings.PROJECT_ROOT, template_dir)
        self.output_dir = os.path.join(settings.PROJECT_ROOT, output_dir)
        self.company_logo = company_logo
        self.default_font = default_font
        
        # Criar diretórios se não existirem
        os.makedirs(self.template_dir, exist_ok=True)
        os.makedirs(self.output_dir, exist_ok=True)
        
        # Configurar ambiente Jinja2 para templates
        self.jinja_env = jinja2.Environment(
            loader=jinja2.FileSystemLoader(self.template_dir),
            autoescape=True
        )
        
        # Configurar pdfkit
        self.pdf_config = configure_pdfkit()
        
        # Configurações padrão para estilos Excel
        self.excel_styles = {
            'header': {
                'font': Font(bold=True, size=12),
                'fill': PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center'),
                'border': Border(
                    bottom=Side(style='thin'),
                    top=Side(style='thin'),
                    left=Side(style='thin'),
                    right=Side(style='thin')
                )
            },
            'data': {
                'font': Font(size=11),
                'alignment': Alignment(horizontal='left', vertical='center'),
                'border': Border(
                    bottom=Side(style='thin'),
                    top=Side(style='thin'),
                    left=Side(style='thin'),
                    right=Side(style='thin')
                )
            }
        }
    
    async def export_to_pdf(
        self,
        data: Union[List[Dict], pd.DataFrame],
        template_name: str,
        output_filename: str,
        title: str,
        subtitle: Optional[str] = None,
        filters: Optional[Dict] = None,
        orientation: str = 'portrait',
        page_size: str = 'A4'
    ) -> str:
        """
        Exporta dados para PDF usando template HTML.
        
        Args:
            data: Dados a serem incluídos no relatório
            template_name: Nome do template HTML
            output_filename: Nome do arquivo PDF de saída
            title: Título do relatório
            subtitle: Subtítulo opcional
            filters: Filtros aplicados aos dados
            orientation: Orientação da página ('portrait' ou 'landscape')
            page_size: Tamanho da página
            
        Returns:
            Caminho do arquivo PDF gerado
        """
        try:
            # Converter DataFrame para lista de dicionários se necessário
            if isinstance(data, pd.DataFrame):
                data = data.to_dict('records')
            
            # Preparar contexto para o template
            context = {
                'title': title,
                'subtitle': subtitle,
                'data': data,
                'filters': filters,
                'company_logo': self.company_logo,
                'generated_at': datetime.datetime.now().strftime('%d/%m/%Y %H:%M'),
                'page_count': '{{ page }} de {{ topage }}',  # Será substituído pelo wkhtmltopdf
                'page_size': page_size,
                'orientation': orientation
            }
            
            # Renderizar template
            template = self.jinja_env.get_template(template_name)
            html_content = template.render(**context)
            
            # Obter opções do PDF
            options = get_pdf_options(
                orientation=orientation,
                page_size=page_size
            )
            
            # Gerar PDF
            output_path = os.path.join(self.output_dir, output_filename)
            pdfkit.from_string(
                html_content,
                output_path,
                options=options,
                configuration=self.pdf_config
            )
            
            logger.info(f"PDF gerado com sucesso: {output_path}")
            return output_path
        
        except Exception as e:
            logger.error(f"Erro ao gerar PDF: {str(e)}")
            raise
    
    async def export_to_excel(
        self,
        data: Union[List[Dict], pd.DataFrame],
        output_filename: str,
        sheet_name: str = 'Relatório',
        title: str = None,
        subtitle: Optional[str] = None,
        filters: Optional[Dict] = None,
        column_widths: Optional[Dict[str, int]] = None,
        freeze_panes: Optional[str] = 'B2'
    ) -> str:
        """
        Exporta dados para Excel com formatação.
        
        Args:
            data: Dados a serem incluídos no relatório
            output_filename: Nome do arquivo Excel de saída
            sheet_name: Nome da planilha
            title: Título do relatório
            subtitle: Subtítulo opcional
            filters: Filtros aplicados aos dados
            column_widths: Larguras personalizadas para colunas
            freeze_panes: Célula para congelar painéis
            
        Returns:
            Caminho do arquivo Excel gerado
        """
        try:
            # Converter para DataFrame se necessário
            if isinstance(data, list):
                df = pd.DataFrame(data)
            else:
                df = data.copy()
            
            # Criar workbook
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Adicionar título e subtítulo se fornecidos
            current_row = 1
            if title:
                ws.cell(row=current_row, column=1, value=title)
                ws.cell(row=current_row, column=1).font = Font(size=14, bold=True)
                current_row += 1
            
            if subtitle:
                ws.cell(row=current_row, column=1, value=subtitle)
                ws.cell(row=current_row, column=1).font = Font(size=12, italic=True)
                current_row += 1
            
            # Adicionar filtros se fornecidos
            if filters:
                current_row += 1
                ws.cell(row=current_row, column=1, value="Filtros aplicados:")
                ws.cell(row=current_row, column=1).font = Font(bold=True)
                current_row += 1
                
                for key, value in filters.items():
                    ws.cell(row=current_row, column=1, value=f"{key}: {value}")
                    current_row += 1
                
                current_row += 1
            
            # Adicionar cabeçalhos
            headers = list(df.columns)
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                self._apply_style(cell, self.excel_styles['header'])
            
            # Adicionar dados
            for row_idx, row in enumerate(df.values, current_row + 1):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    self._apply_style(cell, self.excel_styles['data'])
            
            # Ajustar larguras das colunas
            if column_widths:
                for col, width in column_widths.items():
                    ws.column_dimensions[get_column_letter(col)].width = width
            else:
                self._auto_adjust_columns(ws, headers)
            
            # Congelar painéis
            if freeze_panes:
                ws.freeze_panes = freeze_panes
            
            # Adicionar filtros automáticos
            ws.auto_filter.ref = f"A{current_row}:{get_column_letter(len(headers))}{ws.max_row}"
            
            # Salvar arquivo
            output_path = os.path.join(self.output_dir, output_filename)
            wb.save(output_path)
            
            logger.info(f"Excel gerado com sucesso: {output_path}")
            return output_path
        
        except Exception as e:
            logger.error(f"Erro ao gerar Excel: {str(e)}")
            raise
    
    def _apply_style(self, cell: Any, style: Dict) -> None:
        """
        Aplica estilo a uma célula do Excel.
        
        Args:
            cell: Célula do Excel
            style: Dicionário com estilos a serem aplicados
        """
        for attr, value in style.items():
            setattr(cell, attr, value)
    
    def _auto_adjust_columns(self, worksheet: Any, headers: List[str]) -> None:
        """
        Ajusta automaticamente a largura das colunas.
        
        Args:
            worksheet: Planilha do Excel
            headers: Lista de cabeçalhos
        """
        for col, header in enumerate(headers, 1):
            max_length = len(str(header))
            column_letter = get_column_letter(col)
            
            # Verificar comprimento máximo dos dados na coluna
            for cell in worksheet[column_letter]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Ajustar largura (com um pequeno padding)
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = min(adjusted_width, 50)

# Função para criar instância do exportador com configurações padrão
def create_report_exporter() -> ReportExporter:
    """
    Cria uma instância do exportador de relatórios com configurações padrão.
    
    Returns:
        Instância configurada do ReportExporter
    """
    return ReportExporter(
        company_logo=os.path.join(settings.PROJECT_ROOT, "static/img/logo.png")
    ) 